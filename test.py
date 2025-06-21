import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Load the existing Excel file you just created
file_path = r"C:/Users/sarah.boudarat/PycharmProjects/Thesis/summary_metrics_only.xlsx"
wb = load_workbook(file_path)

for sheet in wb.sheetnames:
    ws = wb[sheet]

    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Format numbers and align columns
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 14  # uniform width

        for cell in col[1:]:  # skip header
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="center")

wb.save(file_path)
print(f"🎓 Professionally formatted Excel saved to: {file_path}")
